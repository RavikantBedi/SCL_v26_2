# from pathlib import Path
# from datetime import datetime

# import polars as pl

# from openpyxl import load_workbook
# from openpyxl.styles import (
#     Font,
#     PatternFill,
#     Border,
#     Side,
#     Alignment
# )


# class ExcelReporter:

#     HEADER_FILL = PatternFill(
#         fill_type="solid",
#         fgColor="1F4E78"
#     )

#     HEADER_FONT = Font(
#         color="FFFFFF",
#         bold=True
#     )

#     THIN_BORDER = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin")
#     )

#     CENTER_ALIGN = Alignment(
#         horizontal="center",
#         vertical="center"
#     )

#     # =====================================
#     # DIRECTORY HANDLING
#     # =====================================

#     def _ensure_dir(self, path: str):
#         Path(path).parent.mkdir(
#             parents=True,
#             exist_ok=True
#         )

#     # =====================================
#     # EXCEL FORMATTING
#     # =====================================

#     def _format_excel(self, file_path: str, sheet_name: str):

#         wb = load_workbook(file_path)
#         ws = wb.active
#         ws.title = sheet_name

#         # Header formatting
#         for cell in ws[1]:
#             cell.fill = self.HEADER_FILL
#             cell.font = self.HEADER_FONT
#             cell.border = self.THIN_BORDER
#             cell.alignment = self.CENTER_ALIGN

#         # Data formatting
#         for row in ws.iter_rows(min_row=2):
#             for cell in row:
#                 cell.border = self.THIN_BORDER
#                 cell.alignment = self.CENTER_ALIGN

#         # Auto column width
#         for column in ws.columns:
#             max_length = 0
#             column_letter = column[0].column_letter

#             for cell in column:
#                 if cell.value is not None:
#                     max_length = max(max_length, len(str(cell.value)))

#             ws.column_dimensions[column_letter].width = max_length + 5

#         # Freeze header
#         ws.freeze_panes = "A2"

#         wb.save(file_path)

#     # =====================================
#     # SAVE EXCEL FILE
#     # =====================================

#     def save_excel(self, df: pl.DataFrame, file_path: str, sheet_name: str):

#         self._ensure_dir(file_path)

#         # IMPORTANT: avoid schema issues
#         df = df.with_columns(pl.all().cast(pl.Utf8))

#         df.write_excel(file_path)

#         self._format_excel(file_path, sheet_name)

#     # =====================================
#     # USER NAME MAPPING
#     # =====================================

#     def _attach_user_name(
#         self,
#         df: pl.DataFrame,
#         user_mapping: pl.DataFrame | None
#     ) -> pl.DataFrame:
#         """
#         Left-joins 'Name' onto df using IP as the key.
#         If user_mapping is None, or the IP has no match,
#         'User Name' is filled with 'Unknown'.

#         Expects:
#           - df has a column called 'IP' (normalized: stripped)
#           - user_mapping has columns 'IP Address' and 'Name'
#             (raw, as read from the 3rd Excel file)
#         """

#         if user_mapping is None or user_mapping.height == 0:
#             return df.with_columns(
#                 pl.lit("Unknown").alias("User Name")
#             )

#         mapping_cmp = user_mapping.select([
#             pl.col("IP Address")
#               .cast(pl.Utf8, strict=False)
#               .fill_null("")
#               .str.strip_chars()
#               .alias("IP"),

#             pl.col("Name")
#               .cast(pl.Utf8, strict=False)
#               .fill_null("Unknown")
#               .str.strip_chars()
#               .alias("User Name"),
#         ])

#         # Drop duplicate IPs in mapping file (keep first occurrence)
#         mapping_cmp = mapping_cmp.unique(subset=["IP"], keep="first")

#         # Drop existing 'User Name' if it exists so we don't get duplicates
#         if "User Name" in df.columns:
#             df = df.drop("User Name")

#         # Left join — preserves all rows of df, adds User Name where found
#         joined = df.join(mapping_cmp, on="IP", how="left")

#         # Fill any unmatched IPs with "Unknown"
#         joined = joined.with_columns(
#             pl.col("User Name").fill_null("Unknown")
#         )

#         # Fallback: if user name is "Unknown" and CompName exists, use CompName instead
#         if "CompName" in joined.columns:
#             joined = joined.with_columns(
#                 pl.when(pl.col("User Name") == "Unknown")
#                 .then(pl.col("CompName").fill_null("Unknown"))
#                 .otherwise(pl.col("User Name"))
#                 .alias("User Name")
#             )

#         return joined

#     # =====================================
#     # REPORT GENERATION
#     # =====================================

#     def generate_reports(
#         self,
#         matched: pl.DataFrame,
#         unmatched: pl.DataFrame,
#         txt_unmatched: pl.DataFrame,
#         inv_unmatched: pl.DataFrame,
#         txt_count: int,
#         user_mapping: pl.DataFrame | None = None,
#         output_dir: str = "output/reports"      # ← NEW: defaults to old behavior
#     ):
#         """
#         output_dir lets the caller (main.py) direct reports into a
#         per-session folder, e.g. output/reports/<session_id>/, so that
#         concurrent users never share or overwrite each other's files.

#         If output_dir is not provided, falls back to the original
#         'output/reports' location for backward compatibility.
#         """

#         base_path = output_dir.rstrip("/") + "/"
#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

#         matched_file = f"{base_path}matched_{timestamp}.xlsx"
#         unmatched_file = f"{base_path}unmatched_combined_{timestamp}.xlsx"
#         txt_unmatched_file = f"{base_path}data_unmatched_{timestamp}.xlsx"
#         inv_unmatched_file = f"{base_path}data_match_{timestamp}.xlsx"
#         summary_file = f"{base_path}summary_{timestamp}.xlsx"

#         # ---------------- ATTACH USER NAME ----------------
#         matched = self._attach_user_name(matched, user_mapping)
#         unmatched = self._attach_user_name(unmatched, user_mapping)
#         txt_unmatched = self._attach_user_name(txt_unmatched, user_mapping)
#         inv_unmatched = self._attach_user_name(inv_unmatched, user_mapping)

#         # ---------------- ALIGN COLUMNS ----------------
#         if "System Model" in matched.columns:
#             matched = matched.drop("System Model")
            
#         if "CompName" in matched.columns:
#             matched = matched.drop("CompName")

#         target_cols = matched.columns
#         for col in target_cols:
#             if col not in unmatched.columns:
#                 unmatched = unmatched.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))
#             if col not in txt_unmatched.columns:
#                 txt_unmatched = txt_unmatched.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))
#             if col not in inv_unmatched.columns:
#                 inv_unmatched = inv_unmatched.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))
        
#         unmatched = unmatched.select(target_cols)
#         txt_unmatched = txt_unmatched.select(target_cols)
#         inv_unmatched = inv_unmatched.select(target_cols)

#         # ---------------- MATCHED ----------------
#         self.save_excel(matched, matched_file, "Matched Assets")

#         # ---------------- UNMATCHED ----------------
#         self.save_excel(unmatched, unmatched_file, "Unmatched Combined")
#         self.save_excel(inv_unmatched, inv_unmatched_file, "Unmatched Category A")
#         self.save_excel(txt_unmatched, txt_unmatched_file, "Unmatched Category B")

#         # ---------------- SUMMARY ----------------
#         match_percentage = (
#             round((matched.height / txt_count) * 100, 2)
#             if txt_count > 0
#             else 0
#         )

#         summary = pl.DataFrame({
#             "Metric": [
#                 "Generated On",
#                 "Total TXT Records",
#                 "Matched Records",
#                 "Unmatched Inventory Records",
#                 "Unmatched Network Records",
#                 "Match Percentage"
#             ],
#             "Value": [
#                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#                 str(txt_count),
#                 str(matched.height),
#                 str(inv_unmatched.height),
#                 str(txt_unmatched.height),
#                 f"{match_percentage}%"
#             ]
#         })

#         self.save_excel(summary, summary_file, "Summary")

#         print("\nReports Generated:")
#         print(matched_file)
#         print(inv_unmatched_file)
#         print(txt_unmatched_file)
#         print(summary_file)


# from pathlib import Path
# from datetime import datetime

# import polars as pl

# from app.core.logger import get_logger

# logger = get_logger()

# from openpyxl import load_workbook
# from openpyxl.styles import (
#     Font,
#     PatternFill,
#     Border,
#     Side,
#     Alignment
# )


# class ExcelReporter:

#     HEADER_FILL = PatternFill(
#         fill_type="solid",
#         fgColor="1F4E78"
#     )

#     HEADER_FONT = Font(
#         color="FFFFFF",
#         bold=True
#     )

#     THIN_BORDER = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin")
#     )

#     CENTER_ALIGN = Alignment(
#         horizontal="center",
#         vertical="center"
#     )

#     # =====================================
#     # DIRECTORY HANDLING
#     # =====================================

#     def _ensure_dir(self, path: str):
#         Path(path).parent.mkdir(
#             parents=True,
#             exist_ok=True
#         )

#     # =====================================
#     # EXCEL FORMATTING
#     # =====================================

#     def _format_excel(self, file_path: str, sheet_name: str):

#         wb = load_workbook(file_path)
#         ws = wb.active
#         ws.title = sheet_name

#         # Header formatting
#         for cell in ws[1]:
#             cell.fill = self.HEADER_FILL
#             cell.font = self.HEADER_FONT
#             cell.border = self.THIN_BORDER
#             cell.alignment = self.CENTER_ALIGN

#         # Data formatting
#         for row in ws.iter_rows(min_row=2):
#             for cell in row:
#                 cell.border = self.THIN_BORDER
#                 cell.alignment = self.CENTER_ALIGN

#         # Auto column width
#         for column in ws.columns:
#             max_length = 0
#             column_letter = column[0].column_letter

#             for cell in column:
#                 if cell.value is not None:
#                     max_length = max(max_length, len(str(cell.value)))

#             ws.column_dimensions[column_letter].width = max_length + 5

#         # Freeze header
#         ws.freeze_panes = "A2"

#         wb.save(file_path)

#     # =====================================
#     # SAVE EXCEL FILE
#     # =====================================

#     def save_excel(self, df: pl.DataFrame, file_path: str, sheet_name: str):

#         self._ensure_dir(file_path)

#         # IMPORTANT: avoid schema issues
#         df = df.with_columns(pl.all().cast(pl.Utf8))

#         df.write_excel(file_path)

#         self._format_excel(file_path, sheet_name)

#     # =====================================
#     # USER NAME MAPPING
#     # =====================================

#     def _attach_user_name(
#         self,
#         df: pl.DataFrame,
#         user_mapping: pl.DataFrame | None
#     ) -> pl.DataFrame:
#         """
#         Left-joins 'Name' onto df using IP as the key.
#         If user_mapping is None, or the IP has no match,
#         'User Name' is filled with 'Unknown'.

#         Expects:
#           - df has a column called 'IP' (normalized: stripped)
#           - user_mapping has columns 'IP Address' and 'Name'
#             (raw, as read from the 3rd Excel file)

#         Duplicate IPs in the mapping file are warned about and
#         deduplicated — the last occurrence wins (most recent entry
#         in the mapping file takes precedence). This prevents the
#         same inconsistency bug as the MAC-only lookup: if two rows
#         in the mapping file have the same IP but different names,
#         we must pick exactly one and do so predictably, not silently
#         pick whichever row happened to appear first.
#         """

#         if user_mapping is None or user_mapping.height == 0:
#             return df.with_columns(
#                 pl.lit("Unknown").alias("User Name")
#             )

#         mapping_cmp = user_mapping.select([
#             pl.col("IP Address")
#               .cast(pl.Utf8, strict=False)
#               .fill_null("")
#               .str.strip_chars()
#               .alias("IP"),

#             pl.col("Name")
#               .cast(pl.Utf8, strict=False)
#               .fill_null("Unknown")
#               .str.strip_chars()
#               .alias("User Name"),
#         ])

#         # Warn and deduplicate if the mapping file has duplicate IPs.
#         # The last occurrence wins (most-recently-listed entry takes priority).
#         # Previously used keep="first" silently — keeping "last" is equally
#         # arbitrary but at least we log the ambiguity so it's visible.
#         duplicate_ips = (
#             mapping_cmp
#             .group_by("IP")
#             .agg(pl.len().alias("count"))
#             .filter(pl.col("count") > 1)
#         )
#         if duplicate_ips.height > 0:
#             dup_list = duplicate_ips.get_column("IP").to_list()
#             logger.warning(
#                 f"User mapping file has {duplicate_ips.height} duplicate IP(s): "
#                 f"{dup_list[:10]}{'...' if len(dup_list) > 10 else ''}. "
#                 f"Last occurrence per IP will be used. "
#                 f"Check the mapping file for inconsistencies."
#             )

#         mapping_cmp = mapping_cmp.unique(subset=["IP"], keep="last")

#         # Drop existing 'User Name' if it exists so we don't get duplicates
#         if "User Name" in df.columns:
#             df = df.drop("User Name")

#         # Left join — preserves all rows of df, adds User Name where found
#         joined = df.join(mapping_cmp, on="IP", how="left")

#         # Fill any unmatched IPs with "Unknown"
#         joined = joined.with_columns(
#             pl.col("User Name").fill_null("Unknown")
#         )

#         # Fallback: if User Name is blank/null/whitespace-only, use CompName instead.
#         # (Previously this only checked for the literal string "Unknown", so a
#         # genuinely blank or whitespace-only name coming from the mapping file
#         # slipped through without ever falling back to CompName.)
#         if "CompName" in joined.columns:
#             is_blank_name = (
#                 pl.col("User Name").is_null()
#                 | (pl.col("User Name").str.strip_chars() == "")
#                 | (pl.col("User Name") == "Unknown")
#             )
#             joined = joined.with_columns(
#                 pl.when(is_blank_name)
#                 .then(
#                     pl.col("CompName")
#                       .cast(pl.Utf8, strict=False)
#                       .fill_null("")
#                       .str.strip_chars()
#                 )
#                 .otherwise(pl.col("User Name"))
#                 .alias("User Name")
#             )
#             # If CompName was also blank/null, fall back to "Unknown" as the final default
#             joined = joined.with_columns(
#                 pl.when(pl.col("User Name").str.strip_chars() == "")
#                 .then(pl.lit("Unknown"))
#                 .otherwise(pl.col("User Name"))
#                 .alias("User Name")
#             )

#         return joined

#     # =====================================
#     # REPORT GENERATION
#     # =====================================

#     def generate_reports(
#         self,
#         matched: pl.DataFrame,
#         unmatched: pl.DataFrame,
#         txt_unmatched: pl.DataFrame,
#         inv_unmatched: pl.DataFrame,
#         txt_count: int,
#         user_mapping: pl.DataFrame | None = None,
#         output_dir: str = "output/reports"
#     ):
#         """
#         output_dir lets the caller (main.py) direct reports into a
#         per-session folder, e.g. output/reports/<session_id>/, so that
#         concurrent users never share or overwrite each other's files.

#         If output_dir is not provided, falls back to the original
#         'output/reports' location for backward compatibility.
#         """

#         base_path = output_dir.rstrip("/") + "/"
#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

#         matched_file = f"{base_path}matched_{timestamp}.xlsx"
#         unmatched_file = f"{base_path}unmatched_combined_{timestamp}.xlsx"
#         txt_unmatched_file = f"{base_path}data_unmatched_{timestamp}.xlsx"
#         inv_unmatched_file = f"{base_path}data_match_{timestamp}.xlsx"
#         summary_file = f"{base_path}summary_{timestamp}.xlsx"

#         # ---------------- DIAGNOSTIC COUNTS (before processing) ----------------
#         logger.info(
#             f"generate_reports received | "
#             f"matched={matched.height}, unmatched={unmatched.height}, "
#             f"category_b(txt_unmatched)={txt_unmatched.height}, "
#             f"category_a(inv_unmatched)={inv_unmatched.height}"
#         )

#         # ---------------- ATTACH USER NAME ----------------
#         matched = self._attach_user_name(matched, user_mapping)
#         unmatched = self._attach_user_name(unmatched, user_mapping)
#         txt_unmatched = self._attach_user_name(txt_unmatched, user_mapping)
#         inv_unmatched = self._attach_user_name(inv_unmatched, user_mapping)

#         # ---------------- ALIGN COLUMNS ----------------
#         # matched: drop internal-only columns before saving
#         if "System Model" in matched.columns:
#             matched = matched.drop("System Model")

#         if "CompName" in matched.columns:
#             matched = matched.drop("CompName")

#         # ── unmatched_combined: union of all cols across the four DataFrames ──
#         # Category A (inv_unmatched) and Category B (txt_unmatched) come from
#         # the combined unmatched pool and may have different column sets from
#         # each other and from matched. We deliberately do NOT force them onto
#         # matched's schema — each file is saved with its own natural columns so
#         # no data is silently dropped.

#         def _drop_internal(df: pl.DataFrame) -> pl.DataFrame:
#             """Remove pipeline-only columns that should never appear in output."""
#             for col in ("System Model", "CompName"):
#                 if col in df.columns:
#                     df = df.drop(col)
#             return df

#         inv_unmatched = _drop_internal(inv_unmatched)
#         txt_unmatched = _drop_internal(txt_unmatched)
#         unmatched     = _drop_internal(unmatched)

#         # For the combined file: build a stable column order (union of all cols,
#         # matched schema first, then any extra cols only in the unmatched pool).
#         target_cols = matched.columns
#         extra_cols = [c for c in unmatched.columns if c not in target_cols]
#         combined_cols = target_cols + extra_cols

#         for col in combined_cols:
#             if col not in unmatched.columns:
#                 unmatched = unmatched.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))

#         unmatched = unmatched.select(combined_cols)

#         # ---------------- MATCHED ----------------
#         self.save_excel(matched, matched_file, "Matched Assets")

#         # ---------------- UNMATCHED ----------------
#         self.save_excel(unmatched, unmatched_file, "Unmatched Combined")
#         self.save_excel(inv_unmatched, inv_unmatched_file, "Unmatched Category A")

#         if "Last AgentCom" in txt_unmatched.columns:
#             txt_unmatched = txt_unmatched.drop("Last AgentCom")

#         self.save_excel(txt_unmatched, txt_unmatched_file, "Unmatched Category B")

#         # ---------------- SUMMARY ----------------
#         match_percentage = (
#             round((matched.height / txt_count) * 100, 2)
#             if txt_count > 0
#             else 0
#         )

#         summary = pl.DataFrame({
#             "Metric": [
#                 "Generated On",
#                 "Total TXT Records",
#                 "Matched Records",
#                 "Unmatched Inventory Records",
#                 "Unmatched Network Records",
#                 "Match Percentage"
#             ],
#             "Value": [
#                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#                 str(txt_count),
#                 str(matched.height),
#                 str(inv_unmatched.height),
#                 str(txt_unmatched.height),
#                 f"{match_percentage}%"
#             ]
#         })

#         self.save_excel(summary, summary_file, "Summary")

#         print("\nReports Generated:")
#         print(matched_file)
#         print(inv_unmatched_file)
#         print(txt_unmatched_file)
#         print(summary_file)














from pathlib import Path
from datetime import datetime

import polars as pl

from app.core.logger import get_logger

logger = get_logger()

from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)


class ExcelReporter:

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    HEADER_FONT = Font(
        color="FFFFFF",
        bold=True
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    CENTER_ALIGN = Alignment(
        horizontal="center",
        vertical="center"
    )

    # IP prefixes to exclude from all unmatched reports.
    # Add more prefixes here as needed — no code changes required elsewhere.
    EXCLUDED_IP_PREFIXES = [
        "192.168.0.",
    ]

    # MAC prefixes to exclude (lowercase, no separators).
    # The filter normalises MAC addresses before comparing, so any separator
    # style (colons, hyphens, dots) in the source data is handled correctly.
    EXCLUDED_MAC_PREFIXES = [
        "7cd30a",
    ]

    # =====================================
    # DIRECTORY HANDLING
    # =====================================

    def _ensure_dir(self, path: str):
        Path(path).parent.mkdir(
            parents=True,
            exist_ok=True
        )

    # =====================================
    # EXCEL FORMATTING
    # =====================================

    def _format_excel(self, file_path: str, sheet_name: str):

        wb = load_workbook(file_path)
        ws = wb.active
        ws.title = sheet_name

        # Header formatting
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
            cell.alignment = self.CENTER_ALIGN

        # Data formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = self.THIN_BORDER
                cell.alignment = self.CENTER_ALIGN

        # Auto column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_length + 5

        # Freeze header
        ws.freeze_panes = "A2"

        wb.save(file_path)

    # =====================================
    # SAVE EXCEL FILE
    # =====================================

    def save_excel(self, df: pl.DataFrame, file_path: str, sheet_name: str):

        self._ensure_dir(file_path)

        # IMPORTANT: avoid schema issues
        df = df.with_columns(pl.all().cast(pl.Utf8))

        df.write_excel(file_path)

        self._format_excel(file_path, sheet_name)

    # =====================================
    # EXCLUSION FILTER
    # =====================================

    def _apply_exclusion_filter(
        self,
        df: pl.DataFrame,
        label: str,
    ) -> tuple[pl.DataFrame, pl.DataFrame]:
        """
        Splits df into two parts:
          - clean:    rows that do NOT match any exclusion rule
          - excluded: rows that DO match at least one exclusion rule

        Exclusion rules (configured via class-level constants):
          1. IP column starts with any prefix in EXCLUDED_IP_PREFIXES
             e.g. "192.168.0." matches 192.165.0.1, 192.165.0.255, etc.
          2. MAC column, normalised to lowercase hex digits only,
             starts with any prefix in EXCLUDED_MAC_PREFIXES
             e.g. "7cd30a" matches 7c:d3:0a:xx:xx:xx in any separator style

        Returns (clean_df, excluded_df).
        If neither IP nor MAC column exists the entire df is returned
        as clean with a warning logged.
        """

        has_ip  = "IP"  in df.columns
        has_mac = "MAC" in df.columns

        if not has_ip and not has_mac:
            logger.warning(
                f"Exclusion filter '{label}': no IP or MAC column found — "
                f"skipping, all {df.height} rows kept."
            )
            return df, df.clear()

        # Build exclusion mask: start with all-False, OR in each rule
        mask = pl.lit(False)

        if has_ip:
            for prefix in self.EXCLUDED_IP_PREFIXES:
                mask = mask | (
                    pl.col("IP")
                      .cast(pl.Utf8, strict=False)
                      .fill_null("")
                      .str.strip_chars()
                      .str.starts_with(prefix)
                )

        if has_mac:
            # Normalise MAC to lowercase hex only before prefix check
            mac_norm = (
                pl.col("MAC")
                  .cast(pl.Utf8, strict=False)
                  .fill_null("")
                  .str.to_lowercase()
                  .str.replace_all(r"[^0-9a-f]", "")
            )
            for prefix in self.EXCLUDED_MAC_PREFIXES:
                mask = mask | mac_norm.str.starts_with(prefix)

        df_with_flag = df.with_columns(mask.alias("_excluded"))

        clean    = df_with_flag.filter(~pl.col("_excluded")).drop("_excluded")
        excluded = df_with_flag.filter( pl.col("_excluded")).drop("_excluded")

        logger.info(
            f"Exclusion filter '{label}': "
            f"{df.height} rows → {clean.height} kept, {excluded.height} excluded"
        )

        return clean, excluded

    # =====================================
    # USER NAME MAPPING
    # =====================================

    def _attach_user_name(
        self,
        df: pl.DataFrame,
        user_mapping: pl.DataFrame | None
    ) -> pl.DataFrame:
        """
        Left-joins 'Name' onto df using IP as the key.
        If user_mapping is None, or the IP has no match,
        'User Name' is filled with 'Unknown'.

        Expects:
          - df has a column called 'IP' (normalized: stripped)
          - user_mapping has columns 'IP Address' and 'Name'

        Duplicate IPs in the mapping file are warned about and
        deduplicated — the last occurrence wins.
        """

        if user_mapping is None or user_mapping.height == 0:
            return df.with_columns(
                pl.lit("Unknown").alias("User Name")
            )

        mapping_cmp = user_mapping.select([
            pl.col("IP Address")
              .cast(pl.Utf8, strict=False)
              .fill_null("")
              .str.strip_chars()
              .alias("IP"),

            pl.col("Name")
              .cast(pl.Utf8, strict=False)
              .fill_null("Unknown")
              .str.strip_chars()
              .alias("User Name"),
        ])

        duplicate_ips = (
            mapping_cmp
            .group_by("IP")
            .agg(pl.len().alias("count"))
            .filter(pl.col("count") > 1)
        )
        if duplicate_ips.height > 0:
            dup_list = duplicate_ips.get_column("IP").to_list()
            logger.warning(
                f"User mapping file has {duplicate_ips.height} duplicate IP(s): "
                f"{dup_list[:10]}{'...' if len(dup_list) > 10 else ''}. "
                f"Last occurrence per IP will be used. "
                f"Check the mapping file for inconsistencies."
            )

        mapping_cmp = mapping_cmp.unique(subset=["IP"], keep="last")

        if "User Name" in df.columns:
            df = df.drop("User Name")

        joined = df.join(mapping_cmp, on="IP", how="left")

        joined = joined.with_columns(
            pl.col("User Name").fill_null("Unknown")
        )

        if "CompName" in joined.columns:
            is_blank_name = (
                pl.col("User Name").is_null()
                | (pl.col("User Name").str.strip_chars() == "")
                | (pl.col("User Name") == "Unknown")
            )
            joined = joined.with_columns(
                pl.when(is_blank_name)
                .then(
                    pl.col("CompName")
                      .cast(pl.Utf8, strict=False)
                      .fill_null("")
                      .str.strip_chars()
                )
                .otherwise(pl.col("User Name"))
                .alias("User Name")
            )
            joined = joined.with_columns(
                pl.when(pl.col("User Name").str.strip_chars() == "")
                .then(pl.lit("Unknown"))
                .otherwise(pl.col("User Name"))
                .alias("User Name")
            )

        return joined

    # =====================================
    # REPORT GENERATION
    # =====================================

    def generate_reports(
        self,
        matched: pl.DataFrame,
        unmatched: pl.DataFrame,
        txt_unmatched: pl.DataFrame,
        inv_unmatched: pl.DataFrame,
        txt_count: int,
        user_mapping: pl.DataFrame | None = None,
        output_dir: str = "output/reports"
    ) -> dict:
        """
        Generates all 5 standard reports plus a 6th filtered_out report
        containing rows removed from the 3 unmatched reports due to the
        IP/MAC exclusion rules defined in EXCLUDED_IP_PREFIXES and
        EXCLUDED_MAC_PREFIXES.

        Returns a dict with:
          - filtered_out_count: rows removed by the IP/MAC exclusion filter
          - unmatched_after_filter: unmatched row count AFTER the exclusion
            filter has been applied (i.e. what is actually written to the
            unmatched.xlsx file), so the dashboard shows the correct number.
        """

        base_path = output_dir.rstrip("/") + "/"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        matched_file       = f"{base_path}matched_{timestamp}.xlsx"
        unmatched_file     = f"{base_path}unmatched_combined_{timestamp}.xlsx"
        txt_unmatched_file = f"{base_path}data_unmatched_{timestamp}.xlsx"
        inv_unmatched_file = f"{base_path}data_match_{timestamp}.xlsx"
        summary_file       = f"{base_path}summary_{timestamp}.xlsx"
        filtered_out_file  = f"{base_path}filtered_out_{timestamp}.xlsx"

        # ---------------- DIAGNOSTIC COUNTS ----------------
        logger.info(
            f"generate_reports received | "
            f"matched={matched.height}, unmatched={unmatched.height}, "
            f"category_b(txt_unmatched)={txt_unmatched.height}, "
            f"category_a(inv_unmatched)={inv_unmatched.height}"
        )

        # ---------------- ATTACH USER NAME ----------------
        matched       = self._attach_user_name(matched, user_mapping)
        unmatched     = self._attach_user_name(unmatched, user_mapping)
        txt_unmatched = self._attach_user_name(txt_unmatched, user_mapping)
        inv_unmatched = self._attach_user_name(inv_unmatched, user_mapping)

        # ---------------- DROP INTERNAL COLUMNS ----------------
        if "System Model" in matched.columns:
            matched = matched.drop("System Model")
        if "CompName" in matched.columns:
            matched = matched.drop("CompName")

        def _drop_internal(df: pl.DataFrame) -> pl.DataFrame:
            for col in ("System Model", "CompName"):
                if col in df.columns:
                    df = df.drop(col)
            return df

        inv_unmatched = _drop_internal(inv_unmatched)
        txt_unmatched = _drop_internal(txt_unmatched)
        unmatched     = _drop_internal(unmatched)

        # ---------------- EXCLUSION FILTER ----------------
        # Applies to all 3 unmatched reports. Only the Unmatched Combined
        # exclusions are written to filtered_out.xlsx to avoid duplicates:
        # Unmatched Combined is already the union of Category A + Category B,
        # so collecting excluded rows from all three would produce 2-3 copies
        # of the same record in filtered_out.
        # The matched report is NOT filtered — matched records are authoritative.

        unmatched,     excl_unmatched     = self._apply_exclusion_filter(unmatched,     "Unmatched Combined")
        inv_unmatched, _excl_inv_unmatched = self._apply_exclusion_filter(inv_unmatched, "Category A")
        txt_unmatched, _excl_txt_unmatched = self._apply_exclusion_filter(txt_unmatched, "Category B")

        # filtered_out.xlsx uses ONLY excl_unmatched (the combined source).
        # Category A / B exclusions are intentionally omitted here because they
        # are already represented in excl_unmatched — using all three would
        # produce duplicate rows in the Filtered Out report.
        if excl_unmatched.height > 0:
            filtered_out_df = excl_unmatched.with_columns(
                pl.lit("Unmatched Combined").alias("Source Report")
            )
        else:
            filtered_out_df = pl.DataFrame({
                "Message": ["No rows were excluded by the IP/MAC filter in this run."]
            })

        filtered_out_count = excl_unmatched.height

        logger.info(
            f"Exclusion filter totals: {filtered_out_count} rows moved to filtered_out "
            f"(from unmatched={excl_unmatched.height}, "
            f"cat_a={_excl_inv_unmatched.height}, "
            f"cat_b={_excl_txt_unmatched.height})"
        )

        # ---------------- BUILD COMBINED UNMATCHED COLUMN ORDER ----------------
        target_cols   = matched.columns
        extra_cols    = [c for c in unmatched.columns if c not in target_cols]
        combined_cols = target_cols + extra_cols

        for col in combined_cols:
            if col not in unmatched.columns:
                unmatched = unmatched.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))

        unmatched = unmatched.select(combined_cols)

        # ---------------- SAVE REPORTS ----------------
        self.save_excel(matched,          matched_file,       "Matched Assets")
        self.save_excel(unmatched,        unmatched_file,     "Unmatched Combined")
        self.save_excel(inv_unmatched,    inv_unmatched_file, "Unmatched Category A")

        if "Last AgentCom" in txt_unmatched.columns:
            txt_unmatched = txt_unmatched.drop("Last AgentCom")

        self.save_excel(txt_unmatched,    txt_unmatched_file, "Unmatched Category B")
        self.save_excel(filtered_out_df,  filtered_out_file,  "Filtered Out Records")

        # ---------------- SUMMARY ----------------
        match_percentage = (
            round((matched.height / txt_count) * 100, 2)
            if txt_count > 0
            else 0
        )

        summary = pl.DataFrame({
            "Metric": [
                "Generated On",
                "Total TXT Records",
                "Matched Records",
                "Unmatched Inventory Records",
                "Unmatched Network Records",
                "Filtered Out Records (IP/MAC exclusion)",
                "Match Percentage",
            ],
            "Value": [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                str(txt_count),
                str(matched.height),
                str(inv_unmatched.height),
                str(txt_unmatched.height),
                str(filtered_out_count),
                f"{match_percentage}%",
            ]
        })

        self.save_excel(summary, summary_file, "Summary")

        print("\nReports Generated:")
        print(matched_file)
        print(inv_unmatched_file)
        print(txt_unmatched_file)
        print(filtered_out_file)
        print(summary_file)

        return {
            "filtered_out_count": filtered_out_count,
            "unmatched_after_filter": unmatched.height,
        }
